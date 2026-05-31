from __future__ import annotations

import os
import re
import smtplib
import ssl
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path

import requests
from fastmcp import FastMCP
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from prettytable import PrettyTable

# 邮箱配置
EMAIL_ADDRESS = "13385816892@163.com"
EMAIL_PASSWORD = "ALxN8bt7ttRQ8bJG"
SMTP_SERVER = "smtp.163.com"
SMTP_PORT = 465

BASE_DIR = Path(__file__).resolve().parent
EXPORT_DIR = BASE_DIR / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

# 初始化（微信客户端延迟加载，便于 Linux/Docker 下仍启动 MCP 其它工具）
mcp = FastMCP("wxauto_mcp")
_wx_client = None
_wx_unavailable_reason: str | None = None


def _get_wechat():
    """Windows + 本机微信可用时返回 WeChat 实例；否则返回 None。"""
    global _wx_client, _wx_unavailable_reason
    if _wx_unavailable_reason is not None:
        return None
    if _wx_client is not None:
        return _wx_client
    try:
        from wxauto4 import WeChat

        _wx_client = WeChat()
        return _wx_client
    except Exception as e:
        _wx_unavailable_reason = str(e)
        return None


@mcp.tool
def add(a: int, b: int) -> int:
    return a + b


@mcp.tool
def hello(name: str = "user") -> str:
    return f"Hello {name}, 这是你的 MCP 服务！"


@mcp.tool
def get_current_time() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S %A")


@mcp.tool
def send_email(to_email: str, content: str) -> str:
    try:
        msg = EmailMessage()
        msg["Subject"] = "来自MCP服务的消息"
        msg["From"] = EMAIL_ADDRESS
        msg["To"] = to_email
        msg.set_content(content)
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, context=context) as smtp:
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            smtp.send_message(msg)
        return f"✅ 邮件已发送到：{to_email}"
    except Exception as e:
        return f"❌ 发送失败：{str(e)}"


# @mcp.tool(name="send_message", description="向微信好友发送消息")
def send_message(msg: str, to: str):
    wx = _get_wechat()
    if wx is None:
        return (
            "❌ 微信不可用（常见于 Docker/Linux 无桌面，或未安装 wxauto4）。"
            f" 原因: {_wx_unavailable_reason or '未知'}"
        )
    try:
        wx.SendMsg(msg, to)
        return "✅ 消息发送成功"
    except Exception as e:
        return f"❌ 发送失败: {str(e)}"


# @mcp.tool(name="get_all_messages", description="获取聊天记录")
# def get_all_messages(who: str):
#     try:
#         wx.ChatWith(who)
#         msgs = wx.GetAllMessage()
#         return [{"sender": m.sender, "content": m.content} for m in msgs if m.type == "friend"]
#     except Exception as e:
#         return f"❌ 获取失败: {str(e)}"


# ---------- 参考 models/output/table_mcp.py + 输出美化.py 的联网与表格能力 ----------


def _safe_excel_filename(name: str | None) -> str:
    base = (name or "").strip() or f"table_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    base = os.path.basename(base)
    if not base.lower().endswith(".xlsx"):
        base += ".xlsx"
    base = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", base).strip(" .")
    if not base or base == ".xlsx":
        base = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return base[:180]


@mcp.tool(
    name="web_search",
    description="联网搜索实时信息（Tavily）。需环境变量 TAVILY_API_KEY；适合新闻、股价、政策等时效问题。",
)
def web_search(query: str, max_results: int = 5) -> str:
    api_key = "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"
    if not api_key:
        return "❌ 未配置 TAVILY_API_KEY：请在运行 mcp_server 的环境中设置该变量后再试。"

    max_results = max(1, min(int(max_results or 5), 10))
    try:
        resp = requests.post(
            "https://api.tavily.com/search",
            json={
                "api_key": api_key,
                "query": query,
                "search_depth": "basic",
                "max_results": max_results,
                "include_answer": True,
            },
            timeout=45,
        )
        resp.raise_for_status()
        data = resp.json()
    except requests.RequestException as e:
        return f"❌ 搜索请求失败：{e}"

    lines: list[str] = []
    ans = data.get("answer")
    if ans:
        lines.append(f"摘要：{ans}")
    for i, item in enumerate(data.get("results") or [], 1):
        title = item.get("title") or ""
        url = item.get("url") or ""
        snippet = (item.get("content") or "")[:500]
        lines.append(f"\n[{i}] {title}\n链接：{url}\n摘录：{snippet}")
    if len(lines) <= (1 if ans else 0):
        return "（未返回有效结果，可尝试改写关键词。）"
    return "\n".join(lines).strip()


@mcp.tool(
    name="format_pretty_table",
    description="根据表头与行数据生成美观表格，支持 markdown/html/latex/text（与 table_mcp 一致）。",
)
def format_pretty_table(
    headers: list,
    rows: list,
    align: str = "c",
    format_type: str = "markdown",
) -> str:
    table = PrettyTable()
    table.field_names = headers
    for row in rows:
        table.add_row(row)
    table.align = align
    table.border = True
    table.header = True

    if format_type == "markdown":
        return str(table)
    if format_type == "html":
        return table.get_html_string()
    if format_type == "latex":
        return table.get_latex_string()
    if format_type == "text":
        return str(table)
    return str(table)


@mcp.tool(
    name="export_to_excel",
    description="将表头与多行数据导出为带样式的 .xlsx 文件，保存到服务目录 exports/ 下并返回绝对路径。",
)
def export_to_excel(headers: list, rows: list, filename: str | None = None) -> str:
    safe_name = _safe_excel_filename(filename)
    out_path = EXPORT_DIR / safe_name

    wb = Workbook()
    ws = wb.active
    ws.title = "数据表格"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="366599")
    center = Alignment(horizontal="center", vertical="center")

    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = center

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    wb.save(str(out_path))
    return f"✅ Excel 已保存：{out_path.resolve()}"


if __name__ == "__main__":
    # _get_wechat()
    # send_message("nh","YU")
    _port = int(os.getenv("MCP_PORT", "8090"))
    mcp.run(transport="http", host="0.0.0.0", port=_port)
